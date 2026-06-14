import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_FOLDER_PATHS = {
    "Foto KTP": r"C:\Users\alfa-raffa\Downloads\(RC) Calon Pelanggan Kegiatan Prioritas Air Minum 2026\Foto KTP (File responses)",
    "Foto Rumah": r"C:\Users\alfa-raffa\Downloads\(RC) Calon Pelanggan Kegiatan Prioritas Air Minum 2026\Foto Rumah (File responses)",
    "Foto Meter Listrik": r"C:\Users\alfa-raffa\Downloads\(RC) Calon Pelanggan Kegiatan Prioritas Air Minum 2026\Foto Meter Listrik (File responses)",
}
FOTO_COLUMNS = ["Foto KTP", "Foto Rumah", "Foto Meter Listrik"]


def extract_filename(path):
    if pd.isna(path):
        return None
    return os.path.basename(str(path)).strip()


def get_matching_sheets(excel_path, required_columns):
    excel_file = pd.ExcelFile(excel_path)
    matching_sheets = []

    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        missing_columns = [col for col in required_columns if col not in df.columns]
        if not missing_columns:
            matching_sheets.append(sheet_name)

    return matching_sheets


def load_sheet_data(excel_path, sheet_name):
    return pd.read_excel(excel_path, sheet_name=sheet_name)


def load_all_sheet_data(excel_path):
    excel_file = pd.ExcelFile(excel_path)
    return {
        sheet_name: pd.read_excel(excel_file, sheet_name=sheet_name)
        for sheet_name in excel_file.sheet_names
    }


def build_file_maps(folder_paths):
    file_maps = {}
    for col, folder_path in folder_paths.items():
        file_maps[col] = {
            file_name.lower(): os.path.join(folder_path, file_name)
            for file_name in os.listdir(folder_path)
        }
    return file_maps


def sync_dataframe(df, files_by_column, sheet_name, foto_columns):
    stats = {
        col: {"filled": 0, "matched": 0, "unmatched": 0}
        for col in foto_columns
    }
    mismatch_details = []

    for idx, row in df.iterrows():
        row_number = idx + 2
        person_name = str(row.get("Nama Lengkap", "")).strip()

        for col in foto_columns:
            filename = extract_filename(row[col])
            lookup_key = filename.lower() if filename else None

            if lookup_key:
                stats[col]["filled"] += 1

            if lookup_key and lookup_key in files_by_column[col]:
                df.at[idx, col] = files_by_column[col][lookup_key]
                stats[col]["matched"] += 1
            else:
                df.at[idx, col] = None
                if lookup_key:
                    stats[col]["unmatched"] += 1
                    mismatch_details.append(
                        {
                            "Sheet": sheet_name,
                            "Baris Excel": row_number,
                            "Nama Lengkap": person_name,
                            "Kolom Foto": col,
                            "Nama File Dicari": filename,
                        }
                    )

    return df, stats, mismatch_details


def build_log_row(sheet_name, stats, foto_columns):
    row = {"Sheet": sheet_name}
    total_matched = 0
    total_unmatched = 0

    for col in foto_columns:
        row[f"{col} Terisi"] = stats[col]["filled"]
        row[f"{col} Cocok"] = stats[col]["matched"]
        row[f"{col} Tidak Cocok"] = stats[col]["unmatched"]
        total_matched += stats[col]["matched"]
        total_unmatched += stats[col]["unmatched"]

    row["Total Cocok"] = total_matched
    row["Total Tidak Cocok"] = total_unmatched
    return row


def build_log_sheet(log_rows):
    return pd.DataFrame(log_rows)


def build_detail_log_sheet(mismatch_details):
    return pd.DataFrame(mismatch_details)


def format_worksheet(worksheet):
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    warning_header_fill = PatternFill(fill_type="solid", fgColor="FDE9D9")
    warning_row_fill = PatternFill(fill_type="solid", fgColor="FDECEC")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    active_header_fill = (
        warning_header_fill if worksheet.title == "Detail Tidak Cocok" else header_fill
    )

    for cell in worksheet[1]:
        cell.fill = active_header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    if worksheet.title == "Detail Tidak Cocok":
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = warning_row_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    if worksheet.title == "Log Sinkronisasi":
        header_map = {
            cell.value: index for index, cell in enumerate(worksheet[1], start=1)
        }
        total_unmatched_col = header_map.get("Total Tidak Cocok")

        if total_unmatched_col:
            for row_index in range(2, worksheet.max_row + 1):
                total_unmatched = (
                    worksheet.cell(row=row_index, column=total_unmatched_col).value or 0
                )
                if total_unmatched > 0:
                    for col_index in range(1, worksheet.max_column + 1):
                        worksheet.cell(
                            row=row_index, column=col_index
                        ).fill = warning_row_fill


def save_workbook(sheets_data, save_path, log_rows=None, mismatch_details=None):
    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets_data.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            format_worksheet(writer.sheets[sheet_name])

        if log_rows:
            build_log_sheet(log_rows).to_excel(
                writer, sheet_name="Log Sinkronisasi", index=False
            )
            format_worksheet(writer.sheets["Log Sinkronisasi"])

        if mismatch_details:
            build_detail_log_sheet(mismatch_details).to_excel(
                writer, sheet_name="Detail Tidak Cocok", index=False
            )
            format_worksheet(writer.sheets["Detail Tidak Cocok"])


def format_counts_summary(stats):
    return ", ".join(
        f"{col}: {values['matched']}" for col, values in stats.items()
    )


def sanitize_filename_part(value):
    cleaned = "".join(
        char if char.isalnum() or char in (" ", "-", "_") else "_"
        for char in value
    ).strip().replace(" ", "_")

    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")

    return cleaned.strip("_")[:40] or "hasil"


def build_output_filename(excel_path, mode_label, sheet_name=None):
    excel_name = os.path.splitext(os.path.basename(excel_path))[0]
    safe_excel_name = sanitize_filename_part(excel_name)
    safe_mode_label = sanitize_filename_part(mode_label)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if sheet_name:
        safe_sheet_name = sanitize_filename_part(sheet_name)
        filename = (
            f"{safe_excel_name}_{safe_sheet_name}_{safe_mode_label}_{timestamp}.xlsx"
        )
    else:
        filename = f"{safe_excel_name}_{safe_mode_label}_{timestamp}.xlsx"

    return os.path.join(os.path.dirname(os.path.abspath(excel_path)), filename)
